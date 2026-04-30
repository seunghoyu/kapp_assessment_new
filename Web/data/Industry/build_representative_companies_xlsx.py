# -*- coding: utf-8 -*-
"""
industry_index.json을 기반으로 소분류(smallName) 단위 산업별 대표기업 3개를 정리하여
industry_representative_companies.xlsx를 생성합니다.
"""

import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Border, Side, Alignment, Font
except ImportError:
    raise SystemExit("openpyxl이 필요합니다. pip install openpyxl")

from industry_representative_companies_data import get_companies

# 경로
BASE = Path(__file__).resolve().parent
INDEX_PATH = BASE / "industry_index.json"
OUTPUT_PATH = BASE / "industry_representative_companies.xlsx"

# 엑셀 컬럼 순서 (기업 1개당 1행, 세로로 읽기)
COLUMNS = [
    "majorName",
    "middleName",
    "smallName",
    "representative_company",
    "company_description",
    "company_type",
    "company_revenue",
    "company_employee_count",
]


def load_unique_industries():
    """industry_index.json에서 (majorName, middleName, smallName) 기준 유일 행 목록 반환."""
    with open(INDEX_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    seen = set()
    rows = []
    for v in data.values():
        key = (v["majorName"], v["middleName"], v["smallName"])
        if key not in seen:
            seen.add(key)
            rows.append({
                "majorName": v["majorName"],
                "middleName": v["middleName"],
                "smallName": v["smallName"],
            })
    return rows


def build_excel_rows(industries):
    """각 산업별 대표기업 3개를 기업 1개당 1행으로 생성 (세로로 읽기)."""
    excel_rows = []
    for ind in industries:
        companies = get_companies(
            ind["majorName"],
            ind["middleName"],
            ind["smallName"],
        )
        # 3개 고정
        while len(companies) < 3:
            companies.append({
                "name": "정보 확인 필요",
                "description": "정보 확인 필요",
                "type": "정보 확인 필요",
                "revenue": "정보 확인 필요",
                "employee_count": "정보 확인 필요",
            })
        for c in companies[:3]:
            excel_rows.append([
                ind["majorName"],
                ind["middleName"],
                ind["smallName"],
                c["name"],
                c["description"],
                c["type"],
                c["revenue"],
                c["employee_count"],
            ])
    return excel_rows


def apply_group_formatting(ws, num_industries):
    """
    소분류(smallName)별 3행 그룹에 대해:
    - majorName, middleName, smallName 열 병합 (3행씩)
    - 그룹 구분선 (굵은 하단 테두리)
    - 병합 셀 세로 가운데 정렬
    """
    thin_side = Side(style="thin")
    thick_bottom = Side(style="medium")  # 그룹 구분용

    for i in range(num_industries):
        start_row = 2 + i * 3
        end_row = 4 + i * 3

        # 1~3열(majorName, middleName, smallName) 병합
        for col in range(1, 4):
            ws.merge_cells(
                start_row=start_row,
                start_column=col,
                end_row=end_row,
                end_column=col,
            )
            cell = ws.cell(row=start_row, column=col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 그룹 마지막 행에 구분선 (하단 굵은 테두리)
        for col in range(1, 9):
            c = ws.cell(row=end_row, column=col)
            c.border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thick_bottom,
            )


def write_xlsx(rows):
    """openpyxl로 XLSX 파일 생성 후 소분류별 병합·구분선 적용."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "대표기업"

    # 헤더
    ws.append(COLUMNS)
    header_font = Font(bold=True)
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=1, column=col).font = header_font

    for row in rows:
        ws.append(row)

    num_industries = len(rows) // 3
    apply_group_formatting(ws, num_industries)

    # 열 너비 조정 (가독성)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    wb.save(OUTPUT_PATH)
    print(f"생성 완료: {OUTPUT_PATH}")


def main():
    industries = load_unique_industries()
    print(f"소분류(unique) 건수: {len(industries)}")
    excel_rows = build_excel_rows(industries)
    write_xlsx(excel_rows)


if __name__ == "__main__":
    main()
